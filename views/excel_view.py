import os
import time
import threading
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side

class ExcelHandler:
    """
    Classe para manipular a geração e salvamento do arquivo Excel
    contendo os dados dos internos.
    """

    def __init__(self, filename: str):
        """
        Parameters
        ----------
        filename : str
            Nome do arquivo Excel (ex: 'Informacoes_Presos-02.xlsx').
        """
        self.filename = filename
        self.wb = Workbook()
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        # Mantém ao menos uma planilha ativa
        self.wb.active

    def create_unit_sheet(self, unit: str, df: pd.DataFrame) -> None:
        """
        Cria uma aba no Excel para a unidade especificada.

        Parameters
        ----------
        unit : str
            Código da unidade (ex. 'PAMC').
        df : pd.DataFrame
            DataFrame contendo as colunas:
            ['Ala', 'Cela', 'Código', 'Preso', 'Mãe', 'Pai', 'Sexo', ...]
        """
        # Remove a planilha padrão "Sheet" se ainda existir
        if "Sheet" in self.wb.sheetnames and len(self.wb.sheetnames) == 1:
            self.wb.remove(self.wb["Sheet"])

        # Cria uma nova aba com o nome da unidade
        ws = self.wb.create_sheet(title=unit)

        # Vamos usar TODAS as colunas do DataFrame na ordem que desejar.
        # Exemplo de ordem manual:
        columns = [
            'Ala',
            'Cela',
            'Código',
            'Preso',
            'Mãe',
            'Pai',
            'Sexo',
            'Data Nasc.',
            'Cidade Origem',
            'Estado',
            'País',
            'Endereço',
            'Estado Civil',
            'Qtd Filhos',
            'Escolaridade',
            'Religião',
            'Profissão',
            'Cor/Etnia',
            'Altura',
            'Modus Operandi',
            'Sentença Dias',
        ]

        # Caso falte alguma coluna, você pode tratar ou criar dinamicamente.
        # Garantir que não quebre se alguma coluna não existir no DF:
        columns = [c for c in columns if c in df.columns]

        # Adiciona o cabeçalho
        ws.append(columns)
        for cell in ws[1]:
            cell.border = self.thin_border

        # Escreve cada linha do DataFrame
        for i, row in df.iterrows():
            data_row = []
            for col in columns:
                data_row.append(row[col])
            ws.append(data_row)

            # Aplica borda nas células da linha recém-adicionada
            for cell in ws[ws.max_row]:
                cell.border = self.thin_border

    def save_periodically(self, interval: int = 300):
        """
        Salva o arquivo Excel periodicamente em segundo plano.
        """
        def auto_save():
            while True:
                time.sleep(interval)
                self.save()

        # Thread que salva a cada N segundos (300 por padrão).
        threading.Thread(target=auto_save, daemon=True).start()

    def save(self):
        """
        Salva o arquivo Excel no disco.
        """
        # Garante ao menos uma sheet visível
        if not any(sheet.sheet_state == 'visible' for sheet in self.wb.worksheets):
            self.wb.active.sheet_state = 'visible'

        temp_filename = self.filename + ".temp"
        self.wb.save(temp_filename)
        os.replace(temp_filename, self.filename)
        print(f"\nArquivo Excel salvo como {self.filename}")
